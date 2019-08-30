from tkinter import *
import openpyxl, threading, os, found_crowd, client, demo1, demo2


def get_message(tk, user_name):
    res = client.main(['2'])
    wb1 = openpyxl.Workbook()
    sheet1 = wb1.active
    sheet1.title = 'sheet1'
    if res[0] == '登录成功':
        for i in res[1]:
            sheet1.append(i)
        wb1.save(r'用户信息\信息.xlsx')
    tk.destroy()
    main(user_name)


def group_chat():
    found_crowd.main()


def fun1(user_name):
    demo2.tcp_server(user_name)


def linkman(tk, f1, user_name, sheet):
    def print_item(event):
        a = mylist.curselection()
        demo1.main([2, user_name, sheet[a[0] + 1][1].value])

    f1.pack_forget()
    f1 = Frame(tk, width='50', height='55')
    f1.place(relx=0, rely=0.2)
    mylist = Listbox(f1, width='47', height='23')
    mylist.bind('<ButtonRelease-1>', print_item)
    for line in range(1, sheet.max_row + 1):
        mylist.insert(END, '用户: ' + sheet[line][2].value)
    mylist.pack(side=LEFT, fill=Y)
    scrollbar = Scrollbar(f1)
    scrollbar.pack(side=RIGHT, fill=Y)
    scrollbar.config(command=mylist.yview)
    mylist.configure(yscrollcommand=scrollbar.set)


def look_group_chat(tk, f1, user_name):
    def print_item(event):
        a = mylist.curselection()
        with open(r'用户信息\群聊\{}'.format(res[a[0]]), 'r', encoding='utf8') as f:
            res_read = f.readlines()
            ip_list = []
            for i in res_read:
                ip_list.append(i[0:-1])
        demo1.main([1, user_name, res[a[0]][0:-4], ip_list])

    f1.pack_forget()

    f1 = Frame(tk, width='50', height='55')
    f1.place(relx=0, rely=0.2)
    res = os.listdir(r'用户信息\群聊')
    mylist = Listbox(f1, width='47', height='23')
    if len(res) == 0:
        mylist.insert(END, '没有群聊')
        mylist['state'] = 'disabled'
    else:
        mylist.bind('<ButtonRelease-1>', print_item)
        for line in range(0, len(res)):
            mylist.insert(END, res[line][0:-4])
    mylist.pack(side=LEFT, fill=Y)
    scrollbar = Scrollbar(f1)
    scrollbar.pack(side=RIGHT, fill=Y)
    scrollbar.config(command=mylist.yview)
    mylist.configure(yscrollcommand=scrollbar.set)


def main(user_name):
    wb = openpyxl.load_workbook(r'用户信息\信息.xlsx')
    sheet = wb['sheet1']

    def print_item(event):
        a = mylist.curselection()
        demo1.main([2, user_name, sheet[a[0] + 1][1].value])

    tk = Tk()
    tk.geometry('350x550')
    tk.title(user_name)
    f1 = Frame(tk, width='50', height='55')
    f1.place(relx=0, rely=0.2)

    mylist = Listbox(f1, width='47', height='23')
    mylist.bind('<ButtonRelease-1>', print_item)
    for line in range(1, sheet.max_row + 1):
        mylist.insert(END, '用户: ' + sheet[line][2].value)
    mylist.pack(side=LEFT, fill=Y)
    scrollbar = Scrollbar(f1)
    scrollbar.pack(side=RIGHT, fill=Y)
    scrollbar.config(command=mylist.yview)
    mylist.configure(yscrollcommand=scrollbar.set)

    btn = Button(tk, text='创建聊天群', command=group_chat)
    btn.place(relx=0.1, rely=0.05)
    btn = Button(tk, text='获取联系人', command=lambda: get_message(tk, user_name))
    btn.place(relx=0.6, rely=0.05)
    btn = Button(tk, text='群聊', width=12, command=lambda: look_group_chat(tk, f1, user_name))
    btn.place(relx=0.6, rely=0.14)
    btn = Button(tk, text='联系人', width=12, command=lambda: linkman(tk, f1, user_name, sheet))
    btn.place(relx=0.1, rely=0.14)
    t1 = threading.Thread(target=fun1, args=(user_name,))
    t1.start()
    tk.mainloop()


if __name__ == '__main__':
    main('人生如梦')
