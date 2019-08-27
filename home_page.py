from tkinter import *
import openpyxl, threading, found_crowd, client
import 聊天客户端 as liao
import 自响应客户端 as die


def get_message(tk, user_name):
    res = client.main(['2'])
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'sheet1'
    if res[0] == '登录成功':
        for i in res[1]:
            sheet.append(i)
        wb.save(r'存储信息\信息.xlsx')
    tk.destroy()
    main(user_name)


def group_chat():
    found_crowd.main()


def fun1(user_name):
    die.tcp_server(user_name)


def main(user_name):
    wb = openpyxl.load_workbook(r'存储信息\信息.xlsx')
    sheet = wb['sheet1']

    def print_item(event):
        a = mylist.curselection()
        liao.main([user_name, sheet[a[0] + 1][1].value])

    tk = Tk()
    tk.geometry('350x550')
    tk.title(user_name)
    f1 = Frame(tk, width='50', height='55')
    f1.place(relx=0, rely=0.2)
    mylist = Listbox(f1, width='47', height='23')
    mylist.bind('<ButtonRelease-1>', print_item)
    for line in range(1, sheet.max_row + 1):
        mylist.insert(END, '用户: ' + sheet[line][2].value)
    mylist.pack(side=LEFT, fill=Y)
    scrollbar = Scrollbar(f1)
    scrollbar.pack(side=RIGHT, fill=Y)
    scrollbar.config(command=mylist.yview)
    mylist.configure(yscrollcommand=scrollbar.set)
    btn = Button(tk, text='创建聊天室', command=lambda: group_chat)
    btn.place(relx=0, rely=0.14)
    btn = Button(tk, text='获取联系人', command=lambda: get_message(tk, user_name))
    btn.place(relx=0.25, rely=0.14)
    t1 = threading.Thread(target=fun1, args=(user_name, ))
    t1.start()
    tk.mainloop()


if __name__ == '__main__':
    main('人生如梦')
